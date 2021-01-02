import time
from shutil import copy2
from PyQt5.QtCore import QThread, pyqtSignal
import numpy as np
import os
import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from openpyxl import load_workbook

temp_path = os.path.join(os.path.expanduser('~'), 'Documents\\Customers')
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

online = 'متصل'
offline = 'غير متصل'
connecting = 'جاري محاولة الاتصال ...'
INTERNET = online

if not os.path.exists(temp_path):
    os.makedirs(temp_path)

if not os.path.exists(os.path.join(temp_path, 'temp.csv')):
    copy2(os.path.join(BASE_DIR, 'Assets\\temp.csv'), temp_path)
    # pd.DataFrame().to_csv(os.path.join(temp_path, 'temp.csv'), index=False, header=False,
    #                       encoding='utf-8-sig')

if not os.path.exists(os.path.join(temp_path, 'cate.csv')):
    copy2(os.path.join(BASE_DIR, 'Assets\\cate.csv'), temp_path)
    # pd.DataFrame().to_csv(os.path.join(temp_path, 'temp.csv'), index=False, header=False,
    #                       encoding='utf-8-sig')

if not os.path.exists(os.path.join(temp_path, 'sales_p.csv')):
    copy2(os.path.join(BASE_DIR, 'Assets\\sales_p.csv'), temp_path)
    # pd.DataFrame().to_csv(os.path.join(temp_path, 'sales_p.csv'), index=False, header=False,
    #                       encoding='utf-8-sig')

if not os.path.exists(os.path.join(temp_path, 'Customer Form.xlsx')):
    copy2(os.path.join(BASE_DIR, 'Assets\\Customer Form.xlsx'), temp_path)


def read_temp():
    customers = pd.read_csv(os.path.join(temp_path, 'temp.csv'), header=0, dtype={
        'phone1': 'str',
        'phone2': 'str',
        'phone3': 'str',
        'phone4': 'str'
    })
    customers = customers.replace(np.nan, '')

    customers['yarn_cate_name'] = customers['yarn_cate_name'].apply((lambda x: np.array(eval(x.replace(' ', ' ,')))))
    customers['omega_cate_name'] = customers['omega_cate_name'].apply((lambda x: np.array(eval(x.replace(' ', ' ,')))))
    customers['factory_cate_name'] = customers['factory_cate_name'].apply(
        (lambda x: np.array(eval(x.replace(' ', ' ,')))))

    cate = pd.read_csv(os.path.join(temp_path, 'cate.csv'), header=0)
    cate = cate.set_index('code')
    cate = cate.replace(np.nan, '')

    sales_p = pd.read_csv(os.path.join(temp_path, 'sales_p.csv'), header=0)
    sales_p = sales_p.set_index('code')
    sales_p = sales_p.replace(np.nan, '')

    return customers, cate, sales_p


class GetAuth(QThread):
    auth = pyqtSignal(object)
    error = pyqtSignal(object)
    off_data = pyqtSignal(object, object, object, object, object, object, str)

    def run(self):
        no_internet = True
        first = 1
        while no_internet:
            try:
                # define the scope
                scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']

                # add credentials to the account
                creds = ServiceAccountCredentials.from_json_keyfile_name('Assets\\token.json', scope)

                # authorize the client sheet
                client = gspread.authorize(creds)
                sheet = client.open('customers')

                self.auth.emit(sheet)
                no_internet = False
                INTERNET = online
            except Exception as e:
                no_internet = True
                INTERNET = offline
                if first:
                    self.error.emit(str(e))
                    first = 0

                cust, cate, sales_p = read_temp()
                edit_requests = pd.DataFrame([], columns=cust.columns)
                self.off_data.emit(cust, cate, sales_p, edit_requests, None, None, INTERNET)

                time.sleep(5)
            print(INTERNET, datetime.time(datetime.now()))


class GetData(QThread):
    sheet = None
    data = pyqtSignal(object, object, object, object, object, object, str)
    error = pyqtSignal(object, object)

    def run(self):
        # get the first sheet of the Spreadsheet
        sheet_customers = self.sheet.get_worksheet(0)
        sheet_category = self.sheet.get_worksheet(1)
        sheet_sales_p = self.sheet.get_worksheet(2)
        sheet_requests = self.sheet.get_worksheet(3)

        while True:
            try:
                # convert the json to dataframe
                customers = pd.DataFrame.from_dict(sheet_customers.get_all_records())
                customers = customers.drop(index=0)
                customers = customers.replace(np.nan, '')

                # print(customers['phone1'])

                edit_requests = pd.DataFrame.from_dict(sheet_requests.get_all_records())
                edit_requests = edit_requests.drop(index=0)
                edit_requests = edit_requests.replace(np.nan, '')

                categories = pd.DataFrame.from_dict(sheet_category.get_all_records())
                categories = categories.set_index('code')

                sales_p = pd.DataFrame.from_dict(sheet_sales_p.get_all_records())
                sales_p = sales_p.set_index('code')

                customers = self.convert_num_to_words(customers, categories)
                edit_requests = self.convert_num_to_words(edit_requests, categories)

                INTERNET = online
                self.data.emit(customers, categories, sales_p, edit_requests, sheet_customers, sheet_requests, INTERNET)

                customers.to_csv(os.path.join(temp_path, 'temp.csv'), index=False, encoding='utf-8-sig')
                categories.to_csv(os.path.join(temp_path, 'cate.csv'), encoding='utf-8-sig')
                sales_p.to_csv(os.path.join(temp_path, 'sales_p.csv'), encoding='utf-8-sig')

            except Exception as e:
                print(e)
                INTERNET = offline
                cust, cate, sales_p = read_temp()
                edit_requests = pd.DataFrame([], columns=cust.columns)
                self.data.emit(cust, cate, sales_p, edit_requests, None, None, INTERNET)

            print(INTERNET, datetime.time(datetime.now()))
            time.sleep(10)

    @staticmethod
    def convert_num_to_words(customers, categories):
        customers = customers.astype({
            'phone1': 'str',
            'phone2': 'str',
            'phone3': 'str',
            'phone4': 'str'
        })

        customers['yarn_cate_name'] = customers['yarn_cate'].apply(lambda x: categories['yarn'].loc[
            list(map(int, x.split(','))) if not isinstance(x, int) and x != '' else [] if x == '' else [
                x]].values)

        customers['omega_cate_name'] = customers['omega_cate'].apply(lambda x: categories['omega'].loc[
            list(map(int, x.split(','))) if not isinstance(x, int) and x != '' else [] if x == '' else [
                x]].values)

        customers['factory_cate_name'] = customers['factory_cate'].apply(lambda x: categories['cloth'].loc[
            list(map(int, x.split(','))) if not isinstance(x, int) and x != '' else [] if x == '' else [
                x]].values)

        return customers


class PrintCustomer(QThread):
    customer = None
    prev_mode = True
    error = pyqtSignal(object)

    def run(self):
        try:
            form_path = os.path.join(temp_path, 'Customer Form.xlsx')

            wb = load_workbook(form_path)

            # grab the active worksheet
            ws = wb.active

            # Data can be assigned directly to cells
            ws['D1'] = self.customer['name'].values[-1]
            ws['E3'] = self.customer['sales_yarn'].values[-1]
            ws['E4'] = self.customer['contact_p'].values[-1]
            ws['E5'] = self.customer['address'].values[-1]
            ws['E6'] = f"0{self.customer['phone1'].values[-1]}" if self.customer['phone1'].values[-1] != '' else ''
            ws['H6'] = f"0{self.customer['phone2'].values[-1]}" if self.customer['phone2'].values[-1] != '' else ''
            ws['K6'] = f"0{self.customer['phone3'].values[-1]}" if self.customer['phone3'].values[-1] != '' else ''
            ws['E7'] = self.customer['e_mail'].values[-1]
            ws['E8'] = self.customer['cust_type'].values[-1]
            ws['E9'] = self.customer['size'].values[-1]
            ws['E10'] = str(self.customer['yarn_cate'].values[-1])
            ws['E11'] = str(self.customer['omega_cate'].values[-1])
            ws['E12'] = str(self.customer['factory_cate'].values[-1])

            # Save the file
            wb.save(form_path)
            if self.prev_mode:
                os.startfile(form_path)
            else:
                os.startfile(form_path, 'print')
        except Exception as e:
            self.error.emit(str(e))


class Save(QThread):
    sheet = None
    sheet_requests = None
    customer = None
    old_customer = None
    auth = None
    auth_type = None
    sheet_row_index = None
    mode = None
    len_data = None

    done = pyqtSignal(str, str)
    error = pyqtSignal(str)

    def run(self):
        try:
            row = [self.customer['i'],  # 0
                   self.customer['sales_yarn'],  # 1
                   self.customer['sales_omega'],  # 2
                   self.customer['sales_cloth'],  # 3
                   self.customer['name'],  # 4
                   self.customer['contact_p'],  # 5
                   self.customer['branch'],  # 6
                   self.customer['area'],  # 7
                   self.customer['address'],  # 8
                   self.customer['phone1'],  # 9
                   self.customer['phone2'],  # 10
                   self.customer['phone3'],  # 11
                   self.customer['phone4'],  # 12
                   self.customer['e_mail'],  # 13
                   self.customer['omega_cate'],  # 14
                   self.customer['yarn_cate'],  # 15
                   self.customer['factory_cate'],  # 16
                   self.customer['size'],  # 17
                   self.customer['factory'],  # 18
                   self.customer['yarn'],  # 19
                   self.customer['omega'],  # 20
                   self.customer['cust_type'],  # 21
                   self.customer['by']  # 22
                   ]

            if self.mode == 'n':
                customers = pd.DataFrame.from_dict(self.sheet.get_all_records())
                customers = customers.drop(index=0)
                customers = customers.replace(np.nan, 0)

                print(type(customers['phone1'].values[-1]))
                try:
                    phone1 = int(row[9])
                except:
                    phone1 = 0

                try:
                    phone2 = int(row[10])
                except:
                    phone2 = 0

                try:
                    phone3 = int(row[11])
                except:
                    phone3 = 0

                try:
                    phone4 = int(row[12])
                except:
                    phone4 = 0

                result = customers[(customers['phone1'] == phone1) |
                                   (customers['phone1'] == phone2) |
                                   (customers['phone1'] == phone3) |
                                   (customers['phone1'] == phone4) |
                                   (customers['phone2'] == phone1) |
                                   (customers['phone2'] == phone2) |
                                   (customers['phone2'] == phone3) |
                                   (customers['phone2'] == phone4) |
                                   (customers['phone3'] == phone1) |
                                   (customers['phone3'] == phone2) |
                                   (customers['phone3'] == phone3) |
                                   (customers['phone3'] == phone4) |
                                   (customers['phone4'] == phone1) |
                                   (customers['phone4'] == phone2) |
                                   (customers['phone4'] == phone3) |
                                   (customers['phone4'] == phone4)]

                if len(result) == 0:
                    row[0] = max(customers['i'].values) + 1
                    row = list(map(str, row))
                    self.sheet.insert_row(row, len(customers) + 3)
                    self.done.emit('نجح', 'تم الاضافه بنجاح')
                else:
                    self.error.emit('هذا العميل موجود من قبل')
            else:
                if self.auth_type == 'admin':
                    for col in range(len(row)):
                        self.sheet.update_cell(self.sheet_row_index, col + 1, str(row[col]))
                    self.done.emit('نجح', 'تم التعديل بنجاح')
                else:
                    edit_requests = pd.DataFrame.from_dict(self.sheet_requests.get_all_records())
                    edit_requests = edit_requests.drop(index=0)
                    edit_requests = edit_requests.replace(np.nan, '')

                    if not self.customer['i'] in edit_requests['i'].values:
                        self.sheet_requests.insert_row(list(map(str, ['new'] + row)), 3)
                        self.sheet_requests.insert_row(list(map(str, ['old'] + [self.old_customer['i'],
                                                                                self.old_customer['sales_yarn'],
                                                                                self.old_customer['sales_omega'],
                                                                                self.old_customer['sales_cloth'],
                                                                                self.old_customer['name'],
                                                                                self.old_customer['contact_p'],
                                                                                self.old_customer['branch'],
                                                                                self.old_customer['area'],
                                                                                self.old_customer['address'],
                                                                                self.old_customer['phone1'],
                                                                                self.old_customer['phone2'],
                                                                                self.old_customer['phone3'],
                                                                                self.old_customer['phone4'],
                                                                                self.old_customer['e_mail'],
                                                                                self.old_customer['omega_cate'],
                                                                                self.old_customer['yarn_cate'],
                                                                                self.old_customer['factory_cate'],
                                                                                self.old_customer['size'],
                                                                                self.old_customer['factory'],
                                                                                self.old_customer['yarn'],
                                                                                self.old_customer['omega'],
                                                                                self.old_customer['cust_type'],
                                                                                self.old_customer['by']])), 3)
                        self.done.emit('', 'تم تقديم طلب للتعديل ...')
                    else:
                        self.error.emit('تم طلب تعديل من قبل لنفس العميل ...')

        except Exception as e:
            self.error.emit(str(e))
